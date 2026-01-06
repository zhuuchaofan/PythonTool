"""
真实端到端测试 - 独立运行版

直接运行此脚本，无需修改 Jcl.py
使用复制的核心类，避免 Jcl.py 模块级初始化问题
"""

import os
import sys
import re
import logging
from collections import defaultdict
from datetime import datetime

# 确保可以导入 openpyxl
try:
    import openpyxl
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(H:%M:%S - %(message)s')
logger = logging.getLogger(__name__)


# ==================== 复制核心类 ====================

class JCLParser:
    """JCL 文件解析器"""
    
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.steps = {}
        self._load_and_parse()

    def _load_and_parse(self):
        try:
            with open(self.filepath, 'r', encoding='utf-8', errors='ignore') as f:
                raw_content = f.read()
            normalized_lines = self._normalize_jcl(raw_content)
            self._parse_lines(normalized_lines)
        except Exception as e:
            print(f"读取 JCL 文件失败: {os.path.basename(self.filepath)} - {e}")

    def _normalize_jcl(self, content: str) -> list:
        lines = content.split('\n')
        cleaned_lines = []
        buffer = ""
        
        for line in lines:
            line = line.strip()
            if not line or line.startswith('//*') or not line.startswith('//'):
                continue
            
            if line.endswith(','):
                if buffer:
                    clean_segment = re.sub(r'^//\s*', '', line)
                    buffer += clean_segment
                else:
                    buffer = line
            else:
                if buffer:
                    clean_segment = re.sub(r'^//\s*', '', line)
                    cleaned_lines.append(buffer + clean_segment)
                    buffer = ""
                else:
                    cleaned_lines.append(line)
        
        return cleaned_lines

    def _parse_lines(self, lines: list):
        current_step_name = None
        re_step = re.compile(r'^//(\S+)\s+EXEC\s+PGM=([A-Z0-9#@$]+)', re.IGNORECASE)
        re_dd = re.compile(r'^//(\S+)\s+DD\s+', re.IGNORECASE)

        for line in lines:
            step_match = re_step.search(line)
            if step_match:
                step_name = step_match.group(1)
                pgm_name = step_match.group(2).upper()
                current_step_name = step_name
                self.steps[step_name] = {"PGM": pgm_name, "DDS": []}
                continue

            if current_step_name:
                dd_match = re_dd.search(line)
                if dd_match:
                    dd_name = dd_match.group(1).upper()
                    dsn = self._extract_param(line, "DSN")
                    if not dsn:
                        continue
                    
                    attrs = {
                        "DD": dd_name,
                        "DSN": dsn,
                        "DISP": self._extract_disp(line),
                        "RECFM": self._extract_param(line, "RECFM"),
                        "LRECL": self._extract_param(line, "LRECL"),
                        "BLKSIZE": self._extract_param(line, "BLKSIZE")
                    }
                    self.steps[current_step_name]["DDS"].append(attrs)

    def _extract_param(self, line: str, key: str) -> str:
        match = re.search(f"{key}=([\\w\\.\\$#@\\(\\)&]+)", line, re.IGNORECASE)
        if match:
            return match.group(1).replace('(', '').replace(')', '')
        return None
    
    def _extract_disp(self, line: str) -> str:
        match = re.search(r'DISP=\(?([A-Z]*)', line, re.IGNORECASE)
        if match:
            disp_val = match.group(1).upper()
            if disp_val in ('NEW', 'OLD', 'SHR', 'MOD'):
                return disp_val
        return None


class AttributeResolver:
    """数据属性推理器"""
    
    SORT_PROGRAMS = {'SORT', 'KQCAMS', 'JEDGENER'}
    
    def __init__(self, group_rows: list):
        self.dsn_map = {r['dataset']: r for r in group_rows if r['dataset']}
    
    def resolve(self, target_dsn: str, jcl_parser: JCLParser) -> tuple:
        if not jcl_parser or not jcl_parser.steps:
            return None, "JCL 中未找到有效的 STEP"

        all_matches = []
        creator_match = None
        
        for step_name, step_data in jcl_parser.steps.items():
            for dd in step_data["DDS"]:
                if dd["DSN"] == target_dsn:
                    all_matches.append((step_name, step_data, dd))
                    if dd.get("DISP") == "NEW" and creator_match is None:
                        creator_match = (step_name, step_data, dd)
        
        if not all_matches:
            return None, "在 JCL 中未找到该 Dataset"
        
        # 优先级 1 & 2: SORT 程序输出
        for step_name, step_data, target_dd in all_matches:
            pgm = step_data["PGM"]
            
            if pgm in self.SORT_PROGRAMS:
                dd_name = target_dd["DD"]
                is_output = dd_name.startswith("SORTOUT") or dd_name == "SYSUT2"
                
                if is_output:
                    meta_info = {"STEP": step_name, "PGM": pgm, "DD": dd_name}
                    
                    if target_dd.get("LRECL") and target_dd.get("RECFM"):
                        return {
                            "Z": "显式定义", "AA": target_dd["RECFM"],
                            "AB": target_dd["LRECL"], "AC": target_dd.get("BLKSIZE", ""),
                            "META": meta_info, "STATUS": "完成(显式)"
                        }, "显式定义"

                    input_dds = [d for d in step_data["DDS"] 
                                if not (d["DD"].startswith("SORTOUT") or d["DD"] == "SYSUT2")]
                    
                    if input_dds:
                        first_input = input_dds[0]
                        source_dsn = first_input["DSN"]
                        
                        if source_dsn in self.dsn_map:
                            src_row = self.dsn_map[source_dsn]
                            return {
                                "Z": source_dsn, "AA": src_row['recfm_val'],
                                "AB": src_row['lrecl_val'], "AC": src_row['blksize_val'],
                                "META": meta_info, "STATUS": "完成(继承)"
                            }, "属性继承"
        
        # 优先级 3: DISP=NEW 创建者
        if creator_match:
            step_name, step_data, target_dd = creator_match
            meta_info = {"STEP": step_name, "PGM": step_data["PGM"], "DD": target_dd["DD"]}
            return {
                "Z": "本JCL创建", "AA": target_dd.get("RECFM"),
                "AB": target_dd.get("LRECL"), "AC": target_dd.get("BLKSIZE"),
                "META": meta_info, "STATUS": "完成(创建)"
            }, "找到创建者"
        
        # 优先级 4: 外部数据集
        step_name, step_data, target_dd = all_matches[0]
        meta_info = {"STEP": step_name, "PGM": step_data["PGM"], "DD": target_dd["DD"]}
        return {
            "Z": "外部数据集", "AA": target_dd.get("RECFM"),
            "AB": target_dd.get("LRECL"), "AC": target_dd.get("BLKSIZE"),
            "META": meta_info, "STATUS": "完成(外部)"
        }, "外部引用"


def build_filename_index(root_dir: str) -> dict:
    """扫描目录，建立文件名索引"""
    file_map = {}
    for root, dirs, files in os.walk(root_dir):
        for file in files:
            name_without_ext = os.path.splitext(file)[0]
            full_path = os.path.join(root, file)
            if name_without_ext not in file_map:
                file_map[name_without_ext] = full_path
    return file_map


# ==================== 测试主函数 ====================

def run_real_test():
    """运行真实的端到端测试"""
    
    print("=" * 60)
    print("JCL 处理程序 - 真实端到端测试")
    print("=" * 60)
    
    TEST_DIR = os.path.dirname(os.path.abspath(__file__))
    TEST_DATA_DIR = os.path.join(TEST_DIR, "test_data")
    JCL_DIR = os.path.join(TEST_DATA_DIR, "JCL")
    EXCEL_FILE = os.path.join(TEST_DATA_DIR, "Test_DSN.xlsx")
    
    if not os.path.exists(EXCEL_FILE):
        print(f"[X] 请先运行 test_integration.py 创建测试文件")
        return False
    
    # 1. 建立 JCL 索引
    print(f"\n[阶段 1] 扫描 JCL 目录...")
    jcl_map = build_filename_index(JCL_DIR)
    print(f"  发现 {len(jcl_map)} 个 JCL 文件: {list(jcl_map.keys())}")
    
    # 2. 读取 Excel
    print(f"\n[阶段 2] 读取 Excel 数据...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Sheet2"]
    
    groups = defaultdict(list)
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        jcl_name = row[2]
        dataset = row[6]
        recfm = row[11]
        lrecl = row[12]
        blksize = row[13]
        
        if not jcl_name:
            continue
        
        recfm_str = str(recfm).strip() if recfm else ""
        needs_process = (recfm_str == "" or recfm_str == "0")
        
        groups[jcl_name].append({
            "row_idx": row_idx, "dataset": dataset,
            "recfm_val": recfm_str, "lrecl_val": lrecl, "blksize_val": blksize,
            "needs_process": needs_process
        })
    
    total_rows = sum(len(v) for v in groups.values())
    print(f"  读取 {total_rows} 条记录, {len(groups)} 个 JCL 分组")
    
    # 3. 处理
    print(f"\n[阶段 3] 解析 JCL 并追溯属性...")
    results = []
    
    for jcl_name, rows in groups.items():
        target_rows = [r for r in rows if r['needs_process']]
        if not target_rows:
            continue
        
        jcl_path = jcl_map.get(jcl_name)
        if not jcl_path:
            print(f"  [!] 找不到 JCL: {jcl_name}")
            continue
        
        print(f"\n  处理 {jcl_name}:")
        parser = JCLParser(jcl_path)
        resolver = AttributeResolver(rows)
        
        for target in target_rows:
            result, status = resolver.resolve(target['dataset'], parser)
            
            if result:
                meta = result.get("META", {})
                print(f"    [OK] {target['dataset']}")
                print(f"       → {result['Z']} | {result['STATUS']}")
                print(f"       → STEP:{meta.get('STEP')} PGM:{meta.get('PGM')} DD:{meta.get('DD')}")
                
                results.append({
                    "row": target['row_idx'],
                    "dataset": target['dataset'],
                    "jcl_name": jcl_name,
                    "result": result
                })
            else:
                print(f"    [X] {target['dataset']} - {status}")
    
    # 4. 写回 Excel
    print(f"\n[阶段 4] 写入结果...")
    
    for item in results:
        row = item["row"]
        r = item["result"]
        meta = r.get("META", {})
        
        ws.cell(row=row, column=26, value=r.get("Z"))
        ws.cell(row=row, column=27, value=r.get("AA"))
        ws.cell(row=row, column=28, value=r.get("AB"))
        ws.cell(row=row, column=29, value=r.get("AC"))
        ws.cell(row=row, column=32, value=r.get("STATUS"))
        ws.cell(row=row, column=33, value=item.get("jcl_name"))
        ws.cell(row=row, column=34, value=meta.get("STEP"))
        ws.cell(row=row, column=35, value=meta.get("PGM"))
        ws.cell(row=row, column=36, value=meta.get("DD"))
    
    output_file = EXCEL_FILE.replace(".xlsx", "_output.xlsx")
    wb.save(output_file)
    print(f"  [OK] 已保存: {output_file}")
    
    # 5. 汇总
    print(f"\n" + "=" * 60)
    print("测试完成!")
    print("=" * 60)
    print(f"  处理记录: {len(results)} 条")
    print(f"  输出文件: {output_file}")
    print(f"\n请打开 Excel 查看 Z~AJ 列的结果")
    
    return True


if __name__ == "__main__":
    run_real_test()
