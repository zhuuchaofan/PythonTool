"""
Excel 相关功能测试

测试场景:
1. validate_excel_structure() 各分支
2. 样例 Excel 文件解析
"""

import os
import sys
import logging
import tempfile

# 使用真实的 openpyxl (不 mock)
import openpyxl
from openpyxl import Workbook

# 配置日志
logging.basicConfig(level=logging.WARNING, format='%(message)s')

# Mock xlwings (只在 Windows 需要) 和 FileHandler
from unittest.mock import MagicMock
sys.modules['xlwings'] = MagicMock()

class MockFileHandler(logging.Handler):
    def __init__(self, *args, **kwargs):
        super().__init__()
    def emit(self, record):
        pass

logging.FileHandler = MockFileHandler

# 引入目标模块
import Jcl


# ==================== 辅助函数 ====================

def create_test_excel(filename: str, sheets: dict, start_row: int = 1) -> str:
    """
    创建测试用 Excel 文件
    
    Args:
        filename: 文件名
        sheets: {sheet_name: [[row1], [row2], ...]}
        start_row: 数据起始行
    
    Returns:
        文件完整路径
    """
    wb = Workbook()
    
    # 删除默认 sheet
    default_sheet = wb.active
    
    for i, (sheet_name, rows) in enumerate(sheets.items()):
        if i == 0:
            ws = default_sheet
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)
        
        for row_idx, row_data in enumerate(rows, start=start_row):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    filepath = os.path.join(tempfile.gettempdir(), filename)
    wb.save(filepath)
    wb.close()
    return filepath


# ==================== validate_excel_structure 测试 ====================

def test_excel_file_not_exist():
    """测试: 文件不存在"""
    print(f"\n{'='*60}")
    print("测试: Excel 文件不存在")
    print(f"{'='*60}")
    
    result, msg = Jcl.validate_excel_structure(
        "/not/exist/file.xlsx", 
        "Sheet1"
    )
    
    if not result and "文件不存在" in msg:
        print(f"  返回: {result}, 消息包含'文件不存在' ✅")
        print(f"\n  🟢 通过")
        return True
    else:
        print(f"  返回: {result}, 消息: {msg} ❌")
        print(f"\n  🔴 失败")
        return False


def test_excel_sheet_not_exist():
    """测试: Sheet 不存在"""
    print(f"\n{'='*60}")
    print("测试: Sheet 不存在")
    print(f"{'='*60}")
    
    # 创建只有 Sheet1 的文件
    filepath = create_test_excel(
        "test_sheet_not_exist.xlsx",
        {"Sheet1": [["A", "B", "C"] + [""] * 40]}  # 至少 40 列
    )
    
    try:
        result, msg = Jcl.validate_excel_structure(filepath, "NotExistSheet")
        
        if not result and "找不到工作表" in msg:
            print(f"  返回: {result}, 消息包含'找不到工作表' ✅")
            print(f"\n  🟢 通过")
            return True
        else:
            print(f"  返回: {result}, 消息: {msg} ❌")
            print(f"\n  🔴 失败")
            return False
    finally:
        os.remove(filepath)


def test_excel_empty_sheet():
    """测试: 空工作表"""
    print(f"\n{'='*60}")
    print("测试: 空工作表")
    print(f"{'='*60}")
    
    # 创建空 sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "EmptySheet"
    filepath = os.path.join(tempfile.gettempdir(), "test_empty_sheet.xlsx")
    wb.save(filepath)
    wb.close()
    
    try:
        result, msg = Jcl.validate_excel_structure(filepath, "EmptySheet")
        
        if not result and "工作表为空" in msg:
            print(f"  返回: {result}, 消息包含'工作表为空' ✅")
            print(f"\n  🟢 通过")
            return True
        else:
            print(f"  返回: {result}, 消息: {msg} ❌")
            print(f"\n  🔴 失败")
            return False
    finally:
        os.remove(filepath)


def test_excel_insufficient_columns():
    """测试: 列数不足"""
    print(f"\n{'='*60}")
    print("测试: 列数不足")
    print(f"{'='*60}")
    
    # 只有 5 列 (需要至少 14 列)
    filepath = create_test_excel(
        "test_few_cols.xlsx",
        {"Sheet1": [["A", "B", "C", "D", "E"]]}
    )
    
    try:
        result, msg = Jcl.validate_excel_structure(filepath, "Sheet1")
        
        if not result and "列数不足" in msg:
            print(f"  返回: {result}, 消息包含'列数不足' ✅")
            print(f"\n  🟢 通过")
            return True
        else:
            print(f"  返回: {result}, 消息: {msg} ❌")
            print(f"\n  🔴 失败")
            return False
    finally:
        os.remove(filepath)


def test_excel_valid_structure():
    """测试: 有效的 Excel 结构"""
    print(f"\n{'='*60}")
    print("测试: 有效的 Excel 结构")
    print(f"{'='*60}")
    
    # 保存原始配置
    original_start_row = Jcl.DATA_START_ROW
    
    # 临时修改起始行为 2 (方便测试)
    Jcl.DATA_START_ROW = 2
    
    # 创建符合要求的 Excel
    # 需要至少 36 列，C 列有 JCL 名，G 列有 Dataset
    header = [""] * 40
    header[0] = "A"
    header[2] = "JCL_NAME"  # C 列
    header[6] = "DATASET"   # G 列
    
    data_row = [""] * 40
    data_row[2] = "TESTJCL"      # C 列: JCL 名
    data_row[6] = "TEST.DATA"    # G 列: Dataset
    
    filepath = create_test_excel(
        "test_valid_excel.xlsx",
        {"Sheet2": [header, data_row]}  # 第 1 行表头，第 2 行数据
    )
    
    try:
        result, msg = Jcl.validate_excel_structure(filepath, "Sheet2")
        
        if result and "验证通过" in msg:
            print(f"  返回: {result} ✅")
            print(f"  验证信息:\n{msg}")
            print(f"\n  🟢 通过")
            return True
        else:
            print(f"  返回: {result}, 消息: {msg} ❌")
            print(f"\n  🔴 失败")
            return False
    finally:
        Jcl.DATA_START_ROW = original_start_row
        os.remove(filepath)


def test_excel_empty_jcl_column():
    """测试: JCL 列全为空"""
    print(f"\n{'='*60}")
    print("测试: JCL 列全为空")
    print(f"{'='*60}")
    
    original_start_row = Jcl.DATA_START_ROW
    Jcl.DATA_START_ROW = 2
    
    # C 列 (JCL) 全空，G 列有数据
    header = [""] * 40
    data_row = [""] * 40
    data_row[6] = "TEST.DATA"  # G 列有数据，但 C 列空
    
    filepath = create_test_excel(
        "test_empty_jcl.xlsx",
        {"Sheet2": [header, data_row]}
    )
    
    try:
        result, msg = Jcl.validate_excel_structure(filepath, "Sheet2")
        
        if not result and "JCL名" in msg and "全部为空" in msg:
            print(f"  返回: {result}, 消息包含 JCL 列为空提示 ✅")
            print(f"\n  🟢 通过")
            return True
        else:
            print(f"  返回: {result}, 消息: {msg} ❌")
            print(f"\n  🔴 失败")
            return False
    finally:
        Jcl.DATA_START_ROW = original_start_row
        os.remove(filepath)


def test_excel_empty_dataset_column():
    """测试: Dataset 列全为空"""
    print(f"\n{'='*60}")
    print("测试: Dataset 列全为空")
    print(f"{'='*60}")
    
    original_start_row = Jcl.DATA_START_ROW
    Jcl.DATA_START_ROW = 2
    
    # C 列有数据，G 列 (Dataset) 全空
    header = [""] * 40
    data_row = [""] * 40
    data_row[2] = "TESTJCL"  # C 列有数据，但 G 列空
    
    filepath = create_test_excel(
        "test_empty_dataset.xlsx",
        {"Sheet2": [header, data_row]}
    )
    
    try:
        result, msg = Jcl.validate_excel_structure(filepath, "Sheet2")
        
        if not result and "Dataset" in msg and "全部为空" in msg:
            print(f"  返回: {result}, 消息包含 Dataset 列为空提示 ✅")
            print(f"\n  🟢 通过")
            return True
        else:
            print(f"  返回: {result}, 消息: {msg} ❌")
            print(f"\n  🔴 失败")
            return False
    finally:
        Jcl.DATA_START_ROW = original_start_row
        os.remove(filepath)


# ==================== 主函数 ====================

def main():
    print("="*60)
    print("Excel 相关功能测试")
    print("="*60)
    
    tests = [
        ("Excel 验证测试", [
            test_excel_file_not_exist,
            test_excel_sheet_not_exist,
            test_excel_empty_sheet,
            test_excel_insufficient_columns,
            test_excel_valid_structure,
            test_excel_empty_jcl_column,
            test_excel_empty_dataset_column,
        ]),
    ]
    
    all_results = []
    
    for category, test_list in tests:
        print(f"\n{'#'*60}")
        print(f"# {category}")
        print(f"{'#'*60}")
        
        for test in test_list:
            try:
                all_results.append(test())
            except Exception as e:
                print(f"\n  💥 异常: {e}")
                import traceback
                traceback.print_exc()
                all_results.append(False)
    
    # 汇总
    print(f"\n{'='*60}")
    print("测试汇总")
    print(f"{'='*60}")
    passed = sum(all_results)
    total = len(all_results)
    print(f"  通过: {passed}/{total}")
    
    if passed == total:
        print("\n  🎉 全部测试通过!")
    else:
        print(f"\n  ⚠️ {total - passed} 个测试失败")
    
    return passed == total


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
