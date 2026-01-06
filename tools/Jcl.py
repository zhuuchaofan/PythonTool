"""
JCL 数据属性解析工具

功能说明:
    从 Excel 文件中读取 Dataset (数据集) 信息，
    通过解析对应的 JCL 文件，自动补全数据的物理属性 (RECFM/LRECL/BLKSIZE)，
    并追溯数据来源，将结果回填到 Excel 中。

适用场景:
    Mainframe 环境下的数据资产梳理。

运行环境:
    Windows (需要 Excel 和 xlwings)
"""

import os
import re
import shutil
import openpyxl
import logging
import time
from collections import defaultdict
from datetime import datetime
import xlwings as xw


# ==================== 配置区域 ====================
# 请根据实际情况修改以下路径和参数

# 基础目录
BASE_DIR = r"C:\Users\zhu-chaofan\Downloads"

# JCL 文件所在根目录
JCL_DIR = os.path.join(BASE_DIR, r"JCL\JCL")

# 输入/输出文件名
SOURCE_FILE_NAME = "DSN_Final.xlsx"
OUTPUT_FILE_NAME = f"AssetList_Lineage_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
LOG_FILE_NAME = f"Process_Log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

# 完整路径
SOURCE_PATH = os.path.join(BASE_DIR, SOURCE_FILE_NAME)
TARGET_PATH = os.path.join(BASE_DIR, OUTPUT_FILE_NAME)
LOG_PATH = os.path.join(BASE_DIR, LOG_FILE_NAME)

# Excel 工作表名称
TARGET_SHEET_NAME = "Sheet2"

# 数据起始行 (从第几行开始读取数据，跳过表头)
DATA_START_ROW = 108415

# 每批处理的数据量
BATCH_SIZE = 1000

# --- 输入列定义 (1 表示 A 列, 2 表示 B 列, 以此类推) ---
COL_JCL_NAME = 3   # C列: JCL 文件名
COL_DATASET = 7    # G列: Dataset 名称
COL_RECFM = 12     # L列: 记录格式 (RECFM)
COL_LRECL = 13     # M列: 记录长度 (LRECL)
COL_BLKSIZE = 14   # N列: 块大小 (BLKSIZE)

# --- 输出列定义 ---
COL_OUT_SOURCE = 26   # Z列: 数据来源
COL_OUT_RECFM = 27    # AA列: RECFM
COL_OUT_LRECL = 28    # AB列: LRECL
COL_OUT_BLKSIZE = 29  # AC列: BLKSIZE
COL_OUT_STATUS = 32   # AF列: 处理状态
COL_OUT_JCL = 33      # AG列: JCL 文件名
COL_OUT_STEP = 34     # AH列: STEP 名称
COL_OUT_PGM = 35      # AI列: 程序名
COL_OUT_DD = 36       # AJ列: DD 名称


# ==================== 日志配置 ====================

def setup_logger(log_file_path: str) -> logging.Logger:
    """初始化日志记录器，同时输出到文件和控制台。"""
    logger = logging.getLogger("JCL_Processor")
    logger.setLevel(logging.INFO)
    
    # 清除已有的处理器
    if logger.handlers:
        logger.handlers.clear()
    
    # 文件日志 (详细记录)
    file_handler = logging.FileHandler(log_file_path, mode='w', encoding='utf-8')
    file_handler.setFormatter(logging.Formatter('%(message)s'))
    logger.addHandler(file_handler)
    
    # 控制台日志 (带时间戳)
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(
        logging.Formatter('%(asctime)s - %(message)s', datefmt='%H:%M:%S')
    )
    logger.addHandler(console_handler)
    
    return logger


logger = setup_logger(LOG_PATH)


# ==================== Excel 结构验证 ====================

def validate_excel_structure(file_path: str, sheet_name: str) -> tuple:
    """
    验证 Excel 文件结构是否符合预期。
    
    检查项:
    1. 文件是否存在
    2. 工作表是否存在
    3. 列数是否足够 (至少包含所有需要读取的列)
    4. 起始行是否有数据
    5. 关键列是否有有效数据 (抽样检查)
    
    Args:
        file_path: Excel 文件路径
        sheet_name: 工作表名称
        
    Returns:
        (是否通过, 错误信息或成功信息)
    """
    logger.info("[预检] 验证 Excel 文件结构...")
    
    # 1. 文件存在性检查
    if not os.path.exists(file_path):
        return False, f"文件不存在: {file_path}"
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        return False, f"无法打开 Excel 文件: {e}"
    
    # 2. 工作表存在性检查
    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        wb.close()
        return False, f"找不到工作表 '{sheet_name}'，可用的工作表: {available}"
    
    ws = wb[sheet_name]
    
    # 3. 列数检查
    required_cols = max(COL_JCL_NAME, COL_DATASET, COL_RECFM, COL_LRECL, COL_BLKSIZE)
    output_cols = max(COL_OUT_SOURCE, COL_OUT_STATUS, COL_OUT_DD)
    
    # 读取第一行检查列数
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if first_row is None:
        wb.close()
        return False, "工作表为空"
    
    actual_cols = len(first_row)
    if actual_cols < required_cols:
        wb.close()
        return False, f"列数不足: 需要至少 {required_cols} 列 (到 {chr(64+required_cols)} 列), 实际只有 {actual_cols} 列"
    
    # 4. 起始行数据检查
    start_row_data = None
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=DATA_START_ROW, values_only=True):
        start_row_data = row
        break
    
    if start_row_data is None:
        wb.close()
        return False, f"起始行 {DATA_START_ROW} 没有数据，请检查 DATA_START_ROW 配置"
    
    # 5. 关键列数据抽样检查 (检查前 100 行)
    sample_count = 0
    valid_jcl_count = 0
    valid_dataset_count = 0
    
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=DATA_START_ROW + 99, values_only=True):
        sample_count += 1
        if len(row) >= required_cols:
            if row[COL_JCL_NAME - 1]:
                valid_jcl_count += 1
            if row[COL_DATASET - 1]:
                valid_dataset_count += 1
    
    wb.close()
    
    if sample_count == 0:
        return False, f"从第 {DATA_START_ROW} 行开始没有数据"
    
    if valid_jcl_count == 0:
        return False, f"抽样 {sample_count} 行中，C 列 (JCL名) 全部为空，请检查列配置"
    
    if valid_dataset_count == 0:
        return False, f"抽样 {sample_count} 行中，G 列 (Dataset) 全部为空，请检查列配置"
    
    # 验证通过
    summary = (
        f"验证通过:\n"
        f"  - 工作表: {sheet_name}\n"
        f"  - 总列数: {actual_cols} (需要读取到第 {required_cols} 列, 写入到第 {output_cols} 列)\n"
        f"  - 起始行: {DATA_START_ROW}\n"
        f"  - 抽样 {sample_count} 行: JCL名有效 {valid_jcl_count} 行, Dataset有效 {valid_dataset_count} 行"
    )
    
    return True, summary


# ==================== JCL 文件索引 ====================

def build_filename_index(root_dir: str) -> dict:
    """
    扫描目录，建立 JCL 文件名到路径的映射。
    
    Args:
        root_dir: JCL 文件根目录
        
    Returns:
        字典 {文件名(不含扩展名): 完整路径}
    """
    logger.info(f"正在扫描 JCL 目录: {root_dir}")
    
    file_map = {}
    file_count = 0
    
    for root, dirs, files in os.walk(root_dir):
        for file in files:
            name_without_ext = os.path.splitext(file)[0]
            full_path = os.path.join(root, file)
            
            # 同名文件只保留第一个
            if name_without_ext not in file_map:
                file_map[name_without_ext] = full_path
            file_count += 1
    
    logger.info(f"扫描完成: 共发现 {file_count} 个文件, 建立 {len(file_map)} 个索引")
    return file_map


# ==================== JCL 解析器 ====================

class JCLParser:
    """
    JCL 文件解析器。
    
    解析 JCL 文件中的:
    - STEP (作业步骤): 每个 EXEC PGM=xxx 语句
    - DD (数据定义): 每个 DD 语句中的 DSN、RECFM、LRECL、BLKSIZE
    """
    
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.steps = {}  # 结构: {"步骤名": {"PGM": "程序名", "DDS": [...]}}
        self._load_and_parse()

    def _load_and_parse(self):
        """加载并解析 JCL 文件。"""
        try:
            with open(self.filepath, 'r', encoding='utf-8', errors='ignore') as f:
                raw_content = f.read()
            
            # 预处理: 合并续行
            normalized_lines = self._normalize_jcl(raw_content)
            # 解析语句
            self._parse_lines(normalized_lines)
            
        except Exception as e:
            logger.error(f"读取 JCL 文件失败: {os.path.basename(self.filepath)} - {e}")

    def _normalize_jcl(self, content: str) -> list:
        """
        预处理 JCL 内容:
        1. 去除注释行 (//*开头)
        2. 合并续行 (以逗号结尾的行与下一行合并)
        """
        lines = content.split('\n')
        cleaned_lines = []
        buffer = ""
        
        for line in lines:
            line = line.strip()
            
            # 跳过空行、注释行、非 JCL 语句
            if not line or line.startswith('//*') or not line.startswith('//'):
                continue
            
            # 处理续行 (以逗号结尾)
            if line.endswith(','):
                if buffer:
                    # 去除续行开头的 //
                    clean_segment = re.sub(r'^//\s*', '', line)
                    buffer += clean_segment
                else:
                    buffer = line
            else:
                if buffer:
                    clean_segment = re.sub(r'^//\s*', '', line)
                    cleaned_lines.append(buffer + clean_segment)
                    buffer = ""
                else:
                    cleaned_lines.append(line)
        
        return cleaned_lines

    def _parse_lines(self, lines: list):
        """
        解析 JCL 语句，提取 STEP 和 DD 信息。
        """
        current_step_name = None
        
        # 正则: 匹配 EXEC PGM=xxx
        re_step = re.compile(r'^//(\S+)\s+EXEC\s+PGM=([A-Z0-9#@$]+)', re.IGNORECASE)
        # 正则: 匹配 DD 语句
        re_dd = re.compile(r'^//(\S+)\s+DD\s+', re.IGNORECASE)

        for line in lines:
            # 识别 STEP (EXEC PGM=xxx)
            step_match = re_step.search(line)
            if step_match:
                step_name = step_match.group(1)
                pgm_name = step_match.group(2).upper()
                
                current_step_name = step_name
                self.steps[step_name] = {
                    "PGM": pgm_name,
                    "DDS": []
                }
                continue

            # 识别 DD 语句 (必须在某个 STEP 内)
            if current_step_name:
                dd_match = re_dd.search(line)
                if dd_match:
                    dd_name = dd_match.group(1).upper()
                    dsn = self._extract_param(line, "DSN")
                    
                    # 没有 DSN 的 DD 跳过
                    if not dsn:
                        continue
                    
                    attrs = {
                        "DD": dd_name,
                        "DSN": dsn,
                        "RECFM": self._extract_param(line, "RECFM"),
                        "LRECL": self._extract_param(line, "LRECL"),
                        "BLKSIZE": self._extract_param(line, "BLKSIZE")
                    }
                    self.steps[current_step_name]["DDS"].append(attrs)

    def _extract_param(self, line: str, key: str) -> str:
        """从 JCL 语句中提取指定参数的值。"""
        match = re.search(f"{key}=([\\w\\.\\$#@\\(\\)]+)", line, re.IGNORECASE)
        if match:
            # 去除括号
            return match.group(1).replace('(', '').replace(')', '')
        return None


# ==================== 血缘推理引擎 ====================

class AttributeResolver:
    """
    数据属性推理器。
    
    根据 JCL 中的信息推导 Dataset 的物理属性:
    1. 如果 DD 中显式定义了属性，直接使用
    2. 如果是 SORT 程序的输出，尝试从输入继承属性
    3. 否则记录为"仅引用"
    """
    
    # 常见的 SORT 类程序
    SORT_PROGRAMS = {'SORT', 'ICEMAN', 'DFSORT', 'SYNCSORT', 'IEBGENER', 'ICEGENER'}
    
    def __init__(self, group_rows: list):
        # 建立 DSN -> Excel行数据 的映射，用于血缘继承
        self.dsn_map = {r['dataset']: r for r in group_rows if r['dataset']}
    
    def resolve(self, target_dsn: str, jcl_parser: JCLParser) -> tuple:
        """
        推导指定 Dataset 的物理属性。
        
        Args:
            target_dsn: 目标 Dataset 名称
            jcl_parser: 已解析的 JCL 对象
            
        Returns:
            (结果字典, 状态描述) 或 (None, 错误描述)
        """
        if not jcl_parser or not jcl_parser.steps:
            return None, "JCL 中未找到有效的 STEP"

        fallback_match = None  # 兜底结果

        for step_name, step_data in jcl_parser.steps.items():
            pgm = step_data["PGM"]
            
            # 在当前 STEP 中查找目标 DSN
            target_dd = next(
                (dd for dd in step_data["DDS"] if dd["DSN"] == target_dsn), 
                None
            )
            
            if not target_dd:
                continue

            # 元数据 (无论什么情况都可以填充)
            meta_info = {
                "STEP": step_name,
                "PGM": pgm,
                "DD": target_dd["DD"]
            }
            
            # 记录兜底结果 (如果没有更好的结果可用)
            if not fallback_match:
                fallback_match = ({
                    "Z": "仅引用",
                    "AA": target_dd["RECFM"],
                    "AB": target_dd["LRECL"],
                    "AC": target_dd["BLKSIZE"],
                    "META": meta_info,
                    "STATUS": "完成(引用)"
                }, "找到引用")

            # 高级逻辑: SORT 程序的输出可以继承输入的属性
            if pgm in self.SORT_PROGRAMS:
                dd_name = target_dd["DD"]
                is_output = dd_name.startswith("SORTOUT") or dd_name == "SYSUT2"
                
                if is_output:
                    # 情况 A: DD 中已显式定义属性
                    if target_dd.get("LRECL") and target_dd.get("RECFM"):
                        return {
                            "Z": "显式定义",
                            "AA": target_dd["RECFM"],
                            "AB": target_dd["LRECL"],
                            "AC": target_dd.get("BLKSIZE", ""),
                            "META": meta_info,
                            "STATUS": "完成(显式)"
                        }, "显式定义"

                    # 情况 B: 从输入 DD 继承属性
                    input_dds = [
                        d for d in step_data["DDS"]
                        if not (d["DD"].startswith("SORTOUT") or d["DD"] == "SYSUT2")
                    ]
                    
                    if input_dds:
                        first_input = input_dds[0]
                        source_dsn = first_input["DSN"]
                        
                        # 检查输入 DSN 是否在 Excel 数据中
                        if source_dsn in self.dsn_map:
                            src_row = self.dsn_map[source_dsn]
                            return {
                                "Z": source_dsn,  # 血缘来源
                                "AA": src_row['recfm_val'],
                                "AB": src_row['lrecl_val'],
                                "AC": src_row['blksize_val'],
                                "META": meta_info,
                                "STATUS": "完成(继承)"
                            }, "属性继承"
        
        # 返回兜底结果
        if fallback_match:
            return fallback_match
            
        return None, "在 JCL 中未找到该 Dataset"


# ==================== 主流程 ====================

def main():
    """主入口函数。"""
    start_time = time.time()
    logger.info(f"========== 任务开始 ==========")
    logger.info(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # ========== 预检: 验证 Excel 结构 ==========
    is_valid, message = validate_excel_structure(SOURCE_PATH, TARGET_SHEET_NAME)
    if not is_valid:
        logger.error(f"[预检失败] {message}")
        return
    logger.info(message)
    
    # 建立 JCL 文件索引
    jcl_path_map = build_filename_index(JCL_DIR)
    
    if not jcl_path_map:
        logger.error("JCL 目录为空或不存在，请检查 JCL_DIR 配置")
        return
    
    # 复制源文件作为输出文件
    logger.info(f"复制文件: {SOURCE_FILE_NAME} -> {OUTPUT_FILE_NAME}")
    shutil.copy2(SOURCE_PATH, TARGET_PATH)

    # ========== 阶段 1: 读取 Excel ==========
    logger.info(f"[阶段 1/3] 读取 Excel 数据 (工作表: {TARGET_SHEET_NAME})")
    
    wb_reader = openpyxl.load_workbook(TARGET_PATH, data_only=True, read_only=True)
    
    try:
        ws_reader = wb_reader[TARGET_SHEET_NAME]
    except KeyError:
        logger.error(f"找不到工作表: {TARGET_SHEET_NAME}")
        return

    # 按 JCL 文件名分组
    groups = defaultdict(list)
    row_counter = 0

    for row in ws_reader.iter_rows(min_row=DATA_START_ROW, values_only=True):
        row_counter += 1
        
        if row_counter % 50000 == 0:
            logger.info(f"  已扫描 {row_counter:,} 行...")
        
        try:
            if len(row) < max(COL_JCL_NAME, COL_DATASET, COL_RECFM):
                continue
            
            jcl_name = row[COL_JCL_NAME - 1]
            if not jcl_name:
                continue
            
            # 检查 RECFM 是否需要处理 (为空或为0)
            recfm_val = row[COL_RECFM - 1]
            recfm_str = str(recfm_val).strip() if recfm_val is not None else ""
            if recfm_str.endswith(".0"):
                recfm_str = recfm_str[:-2]
            
            needs_process = (recfm_str == "0" or recfm_str == "")

            groups[jcl_name].append({
                "row_idx": row_counter + DATA_START_ROW - 1,
                "dataset": row[COL_DATASET - 1],
                "recfm_val": recfm_str,
                "lrecl_val": row[COL_LRECL - 1],
                "blksize_val": row[COL_BLKSIZE - 1],
                "needs_process": needs_process
            })
        except Exception:
            continue
    
    wb_reader.close()
    logger.info(f"  扫描完成: 共 {row_counter:,} 行, {len(groups):,} 个 JCL 分组")

    # ========== 阶段 2: 解析 JCL 并推导血缘 ==========
    logger.info("[阶段 2/3] 解析 JCL 并补全属性")
    
    updates_buffer = []
    jcl_cache = {}
    
    for jcl_name, rows in groups.items():
        # 筛选需要处理的行
        target_rows = [r for r in rows if r['needs_process']]
        if not target_rows:
            continue
        
        # 查找对应的 JCL 文件
        real_path = jcl_path_map.get(jcl_name)
        if not real_path:
            continue

        # 解析 JCL (带缓存)
        if jcl_name not in jcl_cache:
            jcl_cache[jcl_name] = JCLParser(real_path)
        
        parser = jcl_cache[jcl_name]
        resolver = AttributeResolver(rows)

        for target in target_rows:
            result, status = resolver.resolve(target['dataset'], parser)
            
            if result:
                meta = result.get("META", {})
                safe_val = lambda v: v if v else ""
                
                updates_buffer.append({
                    "row": target['row_idx'],
                    # 物理属性 (写入 Z~AC 列)
                    "vals_attr": [
                        safe_val(result["Z"]),   # Z: 数据来源
                        safe_val(result["AA"]),  # AA: RECFM
                        safe_val(result["AB"]),  # AB: LRECL
                        safe_val(result["AC"])   # AC: BLKSIZE
                    ],
                    # 元数据 (写入 AF~AJ 列)
                    "vals_meta": [
                        result.get("STATUS", "完成"),  # AF: 处理状态
                        jcl_name,                       # AG: JCL 文件名
                        safe_val(meta.get("STEP")),     # AH: STEP 名称
                        safe_val(meta.get("PGM")),      # AI: 程序名
                        safe_val(meta.get("DD"))        # AJ: DD 名称
                    ]
                })

    logger.info(f"  解析完成: 共 {len(updates_buffer):,} 条待更新数据")

    # ========== 阶段 3: 回写 Excel ==========
    if updates_buffer:
        total = len(updates_buffer)
        logger.info(f"[阶段 3/3] 回写 Excel (共 {total:,} 条)")
        
        app = xw.App(visible=True)
        app.screen_updating = False
        app.display_alerts = False
        
        try:
            wb = app.books.open(TARGET_PATH)
            app.calculation = 'manual'  # 关闭自动计算以提升性能
            
            try:
                ws = wb.sheets[TARGET_SHEET_NAME]
            except:
                ws = wb.sheets[0]

            for start_idx in range(0, total, BATCH_SIZE):
                end_idx = min(start_idx + BATCH_SIZE, total)
                current_batch = updates_buffer[start_idx:end_idx]
                
                print(f"\n处理进度: {start_idx + 1} ~ {end_idx} / {total}")
                batch_start = time.time()
                
                for i, item in enumerate(current_batch):
                    row_num = item["row"]
                    
                    # 写入物理属性 (Z~AC 列)
                    ws.range((row_num, COL_OUT_SOURCE)).value = item["vals_attr"]
                    
                    # 写入元数据 (AF~AJ 列)
                    ws.range((row_num, COL_OUT_STATUS)).value = item["vals_meta"]
                    
                    if i % 50 == 0:
                        print(f"\r  进度: {i}/{len(current_batch)}", end="")
                
                batch_time = time.time() - batch_start
                print(f"\n  本批耗时: {batch_time:.2f} 秒")
                wb.save()
                
                # 分批确认 (可选)
                if end_idx < total:
                    user_input = input("继续处理下一批? [Y/n]: ").strip().lower()
                    if user_input == 'n':
                        logger.info("用户中断处理")
                        break
        
        except Exception as e:
            logger.error(f"写入 Excel 时发生错误: {e}")
            import traceback
            traceback.print_exc()
        
        finally:
            try:
                app.calculation = 'automatic'
                app.screen_updating = True
                wb.close()
                app.quit()
            except:
                pass
    else:
        logger.info("没有需要更新的数据")

    total_time = time.time() - start_time
    logger.info(f"========== 任务完成 ==========")
    logger.info(f"总耗时: {total_time:.2f} 秒")


if __name__ == "__main__":
    main()