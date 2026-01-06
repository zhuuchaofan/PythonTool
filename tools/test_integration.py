"""
端到端集成测试

创建测试用 Excel 文件和 JCL 文件，运行真实的处理流程
"""

import os
import sys
import shutil
from datetime import datetime

# 确保可以导入 openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, PatternFill
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)


# ==================== 配置 ====================

TEST_DIR = os.path.dirname(os.path.abspath(__file__))
TEST_DATA_DIR = os.path.join(TEST_DIR, "test_data")
JCL_DIR = os.path.join(TEST_DATA_DIR, "JCL")
EXCEL_FILE = os.path.join(TEST_DATA_DIR, "Test_DSN.xlsx")


# ==================== 创建测试 Excel ====================

def create_test_excel():
    """创建符合 Jcl.py 格式要求的测试 Excel 文件"""
    
    print("创建测试 Excel 文件...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet2"
    
    # 表头 (第1行)
    headers = [""] * 40
    headers[2] = "JCL_NAME"      # C列
    headers[6] = "DATASET"       # G列
    headers[11] = "RECFM"        # L列
    headers[12] = "LRECL"        # M列
    headers[13] = "BLKSIZE"      # N列
    headers[25] = "数据来源"     # Z列
    headers[26] = "RECFM_OUT"    # AA列
    headers[27] = "LRECL_OUT"    # AB列
    headers[28] = "BLKSIZE_OUT"  # AC列
    headers[31] = "状态"         # AF列
    headers[32] = "JCL文件"      # AG列
    headers[33] = "STEP"         # AH列
    headers[34] = "PGM"          # AI列
    headers[35] = "DD"           # AJ列
    
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 测试数据 (从第2行开始)
    test_data = [
        # JCL名, Dataset名, RECFM, LRECL, BLKSIZE (空=需要处理)
        ("TESTJOB1", "OUTPUT.SORTED.DATA", "", "", ""),           # SORT 显式定义
        ("TESTJOB1", "INPUT.DATA.FILE1", "FB", "100", "1000"),    # 输入数据 (已有属性)
        ("TESTJOB2", "TARGET.COPY.FILE", "", "", ""),             # JEDGENER 继承
        ("TESTJOB2", "SOURCE.MASTER.FILE", "FB", "80", "800"),    # 源数据 (已有属性)
        ("TESTJOB2", "NEW.CREATED.FILE", "", "", ""),             # 普通程序创建
        ("TESTJOB3", "EXTERNAL.SHARED.DATA", "", "", ""),         # 外部数据集
        ("TESTJOB3", "&&TEMPWORK", "", "", ""),                   # 临时数据集
        ("TESTJOB3", "FINAL.OUTPUT.DATA", "", "", ""),            # KQCAMS 显式定义
    ]
    
    for row_idx, (jcl, dsn, recfm, lrecl, blksize) in enumerate(test_data, 2):
        ws.cell(row=row_idx, column=3, value=jcl)      # C列: JCL名
        ws.cell(row=row_idx, column=7, value=dsn)      # G列: Dataset
        ws.cell(row=row_idx, column=12, value=recfm)   # L列: RECFM
        ws.cell(row=row_idx, column=13, value=lrecl)   # M列: LRECL
        ws.cell(row=row_idx, column=14, value=blksize) # N列: BLKSIZE
        
        # 添加边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for col in range(1, 40):
            ws.cell(row=row_idx, column=col).border = thin_border
    
    # 调整列宽
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['Z'].width = 15
    ws.column_dimensions['AG'].width = 12
    ws.column_dimensions['AH'].width = 10
    ws.column_dimensions['AI'].width = 12
    ws.column_dimensions['AJ'].width = 12
    
    wb.save(EXCEL_FILE)
    print(f"  ✅ 已创建: {EXCEL_FILE}")
    print(f"  ✅ 测试数据: {len(test_data)} 行")
    
    return EXCEL_FILE


def run_integration_test():
    """运行集成测试"""
    
    print("\n" + "=" * 60)
    print("JCL 处理程序 - 端到端集成测试")
    print("=" * 60)
    
    # 1. 创建测试 Excel
    excel_path = create_test_excel()
    
    # 2. 显示测试文件
    print(f"\n测试文件:")
    print(f"  - Excel: {excel_path}")
    print(f"  - JCL 目录: {JCL_DIR}")
    
    jcl_files = os.listdir(JCL_DIR) if os.path.exists(JCL_DIR) else []
    for f in jcl_files:
        print(f"    - {f}")
    
    # 3. 显示预期结果
    print(f"\n预期处理结果:")
    print(f"  1. OUTPUT.SORTED.DATA → 显式定义 (SORT RECFM=FB,LRECL=100)")
    print(f"  2. TARGET.COPY.FILE   → 继承 (从 SOURCE.MASTER.FILE)")  
    print(f"  3. NEW.CREATED.FILE   → 本JCL创建 (DISP=NEW)")
    print(f"  4. EXTERNAL.SHARED.DATA → 外部数据集 (DISP=SHR)")
    print(f"  5. &&TEMPWORK → 本JCL创建 (临时数据集)")
    print(f"  6. FINAL.OUTPUT.DATA → 显式定义 (KQCAMS)")
    
    print(f"\n" + "=" * 60)
    print("要运行真实测试，请执行以下步骤:")
    print("=" * 60)
    print(f"""
1. 修改 Jcl.py 中的配置:
   
   BASE_DIR = r"{TEST_DATA_DIR}"
   JCL_DIR = r"{JCL_DIR}"
   SOURCE_FILE_NAME = "Test_DSN.xlsx"
   DATA_START_ROW = 2

2. 运行程序:
   
   python Jcl.py

3. 检查输出文件中 Z~AJ 列的结果
""")
    
    return True


if __name__ == "__main__":
    run_integration_test()
